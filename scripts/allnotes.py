import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from tkinter import Tk, Button, messagebox, filedialog
import win32com.client

# Supprimer le fichier allnotes_temp.xlsx s'il existe dans le dossier OUTPUT
def delete_temp_excel_file():
    temp_excel_file_path = os.path.abspath('OUTPUT/allnotes_temp.xlsx')
    if os.path.exists(temp_excel_file_path):
        os.remove(temp_excel_file_path)

# Fonction pour sauvegarder le fichier Excel en PDF
def save_allnotes_as_pdf(input_excel, sheet_name, output_pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(input_excel)
    
    # Sélectionner la feuille spécifique
    ws = wb.Worksheets(sheet_name)
    
    # Exporter la feuille en PDF
    ws.ExportAsFixedFormat(0, output_pdf_path)
    
    # Fermer le classeur et quitter Excel
    wb.Close(SaveChanges=False)
    excel.Quit()

def save_selected_allnotes_as_pdf():
    # Déterminer le chemin du fichier Excel et le nom de la feuille
    input_excel = os.path.abspath("OUTPUT/allnotes_temp.xlsx")  # Nom du fichier Excel temporaire
    sheet_name = "allnotes"

    # Ouvrir une boîte de dialogue pour choisir l'emplacement d'enregistrement du fichier PDF
    output_pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
    if output_pdf_path:
        save_allnotes_as_pdf(input_excel, sheet_name, output_pdf_path)

# Fonction principale qui contient votre logique
def telecharger_allnotes():
    try:
        # Supprimer le fichier temporaire s'il existe
        delete_temp_excel_file()

        # Connexion à la base de données SQLite
        conn = sqlite3.connect('data/saisie.db')

        # Récupérer les valeurs depuis la table index
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM 'index'")
        index_row = cursor.fetchone()
        annee_scolaire = index_row[0]
        centre = index_row[1]
        inspecteur = index_row[2]
        membre = index_row[3]

        # Charger les données de la table notes depuis la base de données
        notes_df = pd.read_sql_query("SELECT * FROM notes", conn)

        # Fermer la connexion à la base de données
        conn.close()

        # Charger le fichier Excel existant
        excel_file_path = os.path.abspath('INPUT/allnotes.xlsx')
        if not os.path.exists(excel_file_path):
            raise FileNotFoundError(f"Le fichier {excel_file_path} n'existe pas.")
        
        wb = load_workbook(filename=excel_file_path)
        ws = wb.active

        # Variables pour suivre les changements
        page_global_num = 1
        ligne = 11

        # Charger l'image
        image_path = 'img/logo-tth.png'

        # Fonction pour copier l'en-tête
        def copier_entete(ws: Worksheet, source_row: int, target_row: int, page_num: int):
            for i in range(1, 11):
                for j in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=i, column=j)
                    target_cell = ws.cell(row=target_row + i - 1, column=j)
                    if source_cell.has_style:
                        target_cell._style = source_cell._style
                    target_cell.value = source_cell.value
            # Mettre à jour les cellules spécifiques de l'en-tête
            ws.cell(row=target_row + 2, column=6, value=annee_scolaire)

            # Ajouter l'image dans chaque nouvelle page avec un nom unique
            logo = Image(image_path)
            logo.width = 3.7 * 96  # 9.23 pouces en pixels
            logo.height = 0.9 * 96  # 2.43 pouces en pixels
            logo.anchor = f'B{target_row}'
            ws.add_image(logo)

            # Remplir les cellules spécifiques avec les valeurs récupérées de la table index
            ws.cell(row=target_row + 5, column=6, value=centre)  # F6 de chaque page
            ws.cell(row=target_row + 32, column=3, value=inspecteur)  # C33 de chaque page
            ws.cell(row=target_row + 32, column=8, value=membre)  # H33 de chaque page

        # Fonction pour copier le style d'une ligne
        def copier_style_ligne(ws: Worksheet, source_row: int, target_row: int):
            for j in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=j)
                target_cell = ws.cell(row=target_row, column=j)
                if source_cell.has_style:
                    target_cell._style = source_cell._style

        # Fonction pour calculer le nombre total de pages pour chaque groupe de filtres
        def calculer_total_pages(notes_df):
            total_pages_dict = {}
            for idx, row in notes_df.iterrows():
                key = (row.iloc[3], row.iloc[1], row.iloc[2])
                if key in total_pages_dict:
                    total_pages_dict[key] += 1
                else:
                    total_pages_dict[key] = 1
            for key in total_pages_dict:
                total_pages_dict[key] = (total_pages_dict[key] + 19) // 20  # 20 lignes de données par page
            return total_pages_dict

        total_pages_dict = calculer_total_pages(notes_df)

        copier_entete(ws, 1, 1, page_global_num)

        prev_b7 = notes_df.iloc[0, 3]
        prev_g7 = notes_df.iloc[0, 1]
        prev_k7 = notes_df.iloc[0, 2]
        page_num = 1

        # Insérer les valeurs de b7, g7, et k7 pour la première page
        ws.cell(row=7, column=2, value=prev_b7)
        ws.cell(row=7, column=7, value=prev_g7)
        ws.cell(row=7, column=11, value=prev_k7)

        for idx, row in notes_df.iterrows():
            if ligne > 30 + 37 * (page_global_num - 1) or row.iloc[3] != prev_b7 or row.iloc[1] != prev_g7 or row.iloc[2] != prev_k7:
                if ligne > 11 + 37 * (page_global_num - 1):  # Vérifie si des données ont été ajoutées
                    copier_style_ligne(ws, 32, 32 + 37 * (page_global_num - 1))
                    copier_style_ligne(ws, 33, 33 + 37 * (page_global_num - 1))
                    copier_style_ligne(ws, 37, 37 + 37 * (page_global_num - 1))

                    source_cell = ws.cell(row=32, column=2)
                    target_cell = ws.cell(row=32 + 37 * (page_global_num - 1), column=2)
                    if source_cell.has_style:
                        target_cell._style = source_cell._style
                    target_cell.value = source_cell.value

                    ws.cell(row=33 + 37 * (page_global_num - 1), column=3, value=inspecteur)
                    ws.cell(row=33 + 37 * (page_global_num - 1), column=8, value=membre)

                    total_pages_filter = total_pages_dict[(prev_b7, prev_g7, prev_k7)]
                    ws.cell(row=37 + 37 * (page_global_num - 1), column=6, value=f"{page_num} sur {total_pages_filter}")

                page_global_num += 1
                ligne = 11 + 37 * (page_global_num - 1)
                page_num += 1

                copier_entete(ws, 1, 1 + 37 * (page_global_num - 1), page_global_num)

                ws.cell(row=7 + 37 * (page_global_num - 1), column=2, value=row.iloc[3])
                ws.cell(row=7 + 37 * (page_global_num - 1), column=7, value=row.iloc[1])
                ws.cell(row=7 + 37 * (page_global_num - 1), column=11, value=row.iloc[2])

                if row.iloc[3] != prev_b7 or row.iloc[1] != prev_g7 or row.iloc[2] != prev_k7:
                    page_num = 1

                prev_b7 = row.iloc[3]
                prev_g7 = row.iloc[1]
                prev_k7 = row.iloc[2]

            copier_style_ligne(ws, 11, ligne)
            ws.cell(row=ligne, column=1, value=row.iloc[4])
            ws.cell(row=ligne, column=2, value=row.iloc[5])
            ws.cell(row=ligne, column=3, value=row.iloc[7])
            ws.cell(row=ligne, column=4, value=row.iloc[9])
            ws.cell(row=ligne, column=5, value=row.iloc[10])
            ws.cell(row=ligne, column=6, value=row.iloc[11])
            ws.cell(row=ligne, column=7, value=row.iloc[12])
            ws.cell(row=ligne, column=8, value=row.iloc[13])
            ws.cell(row=ligne, column=9, value=row.iloc[14])
            ws.cell(row=ligne, column=10, value=row.iloc[15])
            ws.cell(row=ligne, column=11, value=row.iloc[16])
            ws.cell(row=ligne, column=12, value=row.iloc[17])

            ligne += 1

        if ligne > 11 + 37 * (page_global_num - 1):  # Vérifie si des données ont été ajoutées
            copier_style_ligne(ws, 32, 32 + 37 * (page_global_num - 1))
            copier_style_ligne(ws, 33, 33 + 37 * (page_global_num - 1))
            copier_style_ligne(ws, 37, 37 + 37 * (page_global_num - 1))

            source_cell = ws.cell(row=32, column=2)
            target_cell = ws.cell(row=32 + 37 * (page_global_num - 1), column=2)
            if source_cell.has_style:
                target_cell._style = source_cell._style
            target_cell.value = source_cell.value

            ws.cell(row=33 + 37 * (page_global_num - 1), column=3, value=inspecteur)
            ws.cell(row=33 + 37 * (page_global_num - 1), column=8, value=membre)

            total_pages_filter = total_pages_dict[(prev_b7, prev_g7, prev_k7)]
            ws.cell(row=37 + 37 * (page_global_num - 1), column=6, value=f"{page_num} sur {total_pages_filter}")

        temp_excel_file_path = os.path.abspath('OUTPUT/allnotes_temp.xlsx')
        wb.save(temp_excel_file_path)

        # Sauvegarde en PDF après la génération du fichier Excel temporaire
        save_selected_allnotes_as_pdf()

        messagebox.showinfo("ممتاز", "تم تصدير محضر النقط بنجاح")

    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur s'est produite: {e}")

# Appel de la fonction telecharger_allnotes() au démarrage du script
if __name__ == "__main__":
    telecharger_allnotes()
