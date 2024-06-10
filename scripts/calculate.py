import os
from pathlib import Path
import xlwings as xw  
import tkinter as tk
from tkinter import messagebox
import sqlite3
import openpyxl
import datetime

def delete_temp_excel_files():
    temp_excel_file_path1 = os.path.abspath('OUTPUT/note_temp.xlsx')
    temp_excel_file_path2 = os.path.abspath('OUTPUT/note.xlsx')
    
    if os.path.exists(temp_excel_file_path1):
        os.remove(temp_excel_file_path1)
    if os.path.exists(temp_excel_file_path2):
        os.remove(temp_excel_file_path2)

delete_temp_excel_files()

def convertir_valeurs(row):
    return [str(cell) if isinstance(cell, (datetime.time, datetime.date, datetime.datetime)) else cell for cell in row]

def remplir_et_convertir_et_importer():
    conn = sqlite3.connect('data/saisie.db')
    c = conn.cursor()

    wb = openpyxl.load_workbook('INPUT/note.xlsx')
    sheet = wb['saisie']  # Assurez-vous que le nom de la feuille est correct

    c.execute("SELECT * FROM saisie")
    donnees = c.fetchall()

    ligne = 2
    colonnes_map = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'O']
    for row in donnees:
        for i, colonne in enumerate(colonnes_map):
            sheet[f"{colonne}{ligne}"] = row[i]
        ligne += 1

    formule_l = sheet['L2'].value
    formule_n = sheet['N2'].value
    formule_p = sheet['P2'].value
    formule_q = sheet['Q2'].value
    formule_r = sheet['R2'].value
    formule_s = sheet['S2'].value
    formule_t = sheet['T2'].value
    formule_u = sheet['U2'].value
    formule_v = sheet['V2'].value

    for row in range(3, sheet.max_row + 1):
        nouvelle_formule_l = formule_l.replace("I2", f"I{row}")
        for cellule in ['J', 'K', 'M']:
            nouvelle_formule_l = nouvelle_formule_l.replace(f"{cellule}2", f"{cellule}{row}")
        sheet[f"L{row}"] = nouvelle_formule_l

        nouvelle_formule_n = formule_n.replace("J2", f"J{row}").replace("M2", f"M{row}").replace("I2", f"I{row}")
        sheet[f"N{row}"] = nouvelle_formule_n

        nouvelle_formule_p = formule_p.replace("J2", f"J{row}").replace("O2", f"O{row}").replace("I2", f"I{row}")
        sheet[f"P{row}"] = nouvelle_formule_p

        nouvelle_formule_q = formule_q.replace("J2", f"J{row}").replace("L2", f"L{row}").replace("N2", f"N{row}").replace("P2", f"P{row}")
        sheet[f"Q{row}"] = nouvelle_formule_q

        nouvelle_formule_r = formule_r.replace("J2", f"J{row}").replace("S2", f"S{row}").replace("L2", f"L{row}").replace("N2", f"N{row}").replace("P2", f"P{row}")
        sheet[f"R{row}"] = nouvelle_formule_r

        nouvelle_formule_s = formule_s.replace("K2", f"K{row}").replace("M2", f"M{row}").replace("O2", f"O{row}").replace("J2", f"J{row}")
        sheet[f"S{row}"] = nouvelle_formule_s

        nouvelle_formule_t = formule_t.replace("R2", f"R{row}")
        sheet[f"T{row}"] = nouvelle_formule_t

        nouvelle_formule_u = formule_u.replace("R2", f"R{row}")
        sheet[f"U{row}"] = nouvelle_formule_u

        nouvelle_formule_v = formule_v.replace("R2", f"R{row}")
        sheet[f"V{row}"] = nouvelle_formule_v

    wb.save('OUTPUT/note_temp.xlsx')

    current_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()
    file_path = current_dir / "OUTPUT/note_temp.xlsx"
    note_path = current_dir / "OUTPUT/note.xlsx"

    with xw.App(visible=False) as app:
        wb_xw = app.books.open(file_path)
        for sheet in wb_xw.sheets:
            sheet.used_range.value = sheet.used_range.value
        wb_xw.save(note_path)
        wb_xw.close()

    wb = openpyxl.load_workbook(note_path)
    sheet = wb['saisie']

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        row = convertir_valeurs(row)
        # Vérifier si l'enregistrement existe déjà dans la base de données
        c.execute('SELECT COUNT(*) FROM notes WHERE massar = ?', (row[5],))
        if c.fetchone()[0] == 0:
            c.execute('INSERT INTO notes VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', row)

    conn.commit()
    conn.close()

    messagebox.showinfo("ممتاز", "تم حساب المعدلات بنجاح")

# Appel de la fonction telecharger_NF() au démarrage du script
if __name__ == "__main__":
    remplir_et_convertir_et_importer()
