import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image
from tkinter import Tk, Button, messagebox, filedialog
import win32com.client
import subprocess
import sys

def resource_path(relative_path):
    """Obtenir le chemin absolu vers la ressource, fonctionne pour dev et pour PyInstaller"""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Supprimer le fichier NF_temp.xlsx s'il existe dans le dossier OUTPUT
def delete_temp_excel_file():
    temp_excel_file_path = resource_path('OUTPUT/NF_temp.xlsx')
    if os.path.exists(temp_excel_file_path):
        os.remove(temp_excel_file_path)

# Fonction pour sauvegarder le fichier Excel en PDF
def save_NF_as_pdf(input_excel, sheet_name, output_pdf_path):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(input_excel)
    
    # Sélectionner la feuille spécifique
    ws = wb.Worksheets(sheet_name)
    
    # Exporter la feuille en PDF
    ws.ExportAsFixedFormat(0, output_pdf_path)
    
    # Fermer le classeur et quitter Excel
    wb.Close(SaveChanges=False)
    excel.Quit()

def save_selected_NF_as_pdf():
    # Déterminer le chemin du fichier Excel et le nom de la feuille
    input_excel = resource_path('OUTPUT/NF_temp.xlsx')  # Nom du fichier Excel temporaire
    sheet_name = "NF"

    # Ouvrir une boîte de dialogue pour choisir l'emplacement d'enregistrement du fichier PDF
    output_pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
    if output_pdf_path:
        save_NF_as_pdf(input_excel, sheet_name, output_pdf_path)

# Fonction principale qui contient votre logique
def telecharger_NF():
    try:
        # Supprimer le fichier temporaire s'il existe
        delete_temp_excel_file()

        # Connexion à la base de données SQLite
        conn = sqlite3.connect('data/saisie.db')

        # Récupérer les valeurs depuis la table index
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM 'index'")
        index_row = cursor.fetchone()
        annee_scolaire = index_row[0]
        centre = index_row[1]
        inspecteur = index_row[2]
        membre = index_row[3]

        # Charger les données de la table notes depuis la base de données
        notes_df = pd.read_sql_query("SELECT * FROM notes", conn)

        # Fermer la connexion à la base de données
        conn.close()

        # Charger le fichier Excel existant
        excel_file_path = resource_path('INPUT/NF.xlsx')

        # Afficher le chemin du fichier pour débogage
        print(f"Chemin du fichier Excel : {excel_file_path}")
        if not os.path.exists(excel_file_path):
            raise FileNotFoundError(f"Le fichier {excel_file_path} n'existe pas.")
        
        wb = load_workbook(filename=excel_file_path)
        ws = wb.active

        # Variables pour suivre les changements
        page_global_num = 1
        ligne = 13

        # Charger l'image
        image_path = resource_path('img/logo-tth.png')
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Le fichier image {image_path} n'existe pas.")

        # Fonction pour copier l'en-tête
        def copier_entete(ws: Worksheet, source_row: int, target_row: int, page_num: int):
            for i in range(1, 13):
                for j in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=i, column=j)
                    target_cell = ws.cell(row=target_row + i - 1, column=j)
                    if source_cell.has_style:
                        target_cell._style = source_cell._style
                    target_cell.value = source_cell.value
            # Mettre à jour les cellules spécifiques de l'en-tête
            ws.cell(row=target_row + 5, column=5, value=annee_scolaire)

            # Ajouter l'image dans chaque nouvelle page avec un nom unique
            logo = Image(image_path)
            logo.width = 3.7 * 96  # 9.23 pouces en pixels
            logo.height = 0.9 * 96  # 2.43 pouces en pixels
            logo.anchor = f'A{target_row}'
            ws.add_image(logo)

            # Remplir les cellules spécifiques avec les valeurs récupérées de la table index
            ws.cell(row=target_row + 8, column=6, value=centre)  # F9 de chaque page
            ws.cell(row=target_row + 39, column=2, value=inspecteur)  # B40 de chaque page
            ws.cell(row=target_row + 39, column=6, value=membre)  # F40 de chaque page

        # Fonction pour copier le style d'une ligne
        def copier_style_ligne(ws: Worksheet, source_row: int, target_row: int):
            for j in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=j)
                target_cell = ws.cell(row=target_row, column=j)
                if source_cell.has_style:
                    target_cell._style = source_cell._style

        # Fonction pour calculer le nombre total de pages pour chaque groupe de filtres
        def calculer_total_pages(notes_df):
            total_pages_dict = {}
            for idx, row in notes_df.iterrows():
                key = (row.iloc[3], row.iloc[1], row.iloc[2])
                if key in total_pages_dict:
                    total_pages_dict[key] += 1
                else:
                    total_pages_dict[key] = 1
            for key in total_pages_dict:
                total_pages_dict[key] = (total_pages_dict[key] + 23) // 24  # 24 lignes de données par page
            return total_pages_dict

        total_pages_dict = calculer_total_pages(notes_df)

        copier_entete(ws, 1, 1, page_global_num)

        prev_b10 = notes_df.iloc[0, 3]
        prev_f10 = notes_df.iloc[0, 1]
        prev_h10 = notes_df.iloc[0, 2]
        page_num = 1

        # Insérer les valeurs de B10, F10, et H10 pour la première page
        ws.cell(row=10, column=2, value=prev_b10)
        ws.cell(row=10, column=6, value=prev_f10)
        ws.cell(row=10, column=8, value=prev_h10)

        for idx, row in notes_df.iterrows():
            if ligne > 36 + 43 * (page_global_num - 1) or row.iloc[3] != prev_b10 or row.iloc[1] != prev_f10 or row.iloc[2] != prev_h10:
                if ligne > 13 + 43 * (page_global_num - 1):  # Vérifie si des données ont été ajoutées
                    copier_style_ligne(ws, 38, 38 + 43 * (page_global_num - 1))
                    copier_style_ligne(ws, 40, 40 + 43 * (page_global_num - 1))
                    copier_style_ligne(ws, 43, 43 + 43 * (page_global_num - 1))

                    source_cell = ws.cell(row=38, column=2)
                    target_cell = ws.cell(row=38 + 43 * (page_global_num - 1), column=2)
                    if source_cell.has_style:
                        target_cell._style = source_cell._style
                    target_cell.value = source_cell.value

                    ws.cell(row=40 + 43 * (page_global_num - 1), column=2, value=inspecteur)
                    ws.cell(row=40 + 43 * (page_global_num - 1), column=6, value=membre)

                    total_pages_filter = total_pages_dict[(prev_b10, prev_f10, prev_h10)]
                    ws.cell(row=43 + 43 * (page_global_num - 1), column=4, value=f"{page_num} sur {total_pages_filter}")

                page_global_num += 1
                ligne = 13 + 43 * (page_global_num - 1)
                page_num += 1

                copier_entete(ws, 1, 1 + 43 * (page_global_num - 1), page_global_num)

                ws.cell(row=10 + 43 * (page_global_num - 1), column=2, value=row.iloc[3])
                ws.cell(row=10 + 43 * (page_global_num - 1), column=6, value=row.iloc[1])
                ws.cell(row=10 + 43 * (page_global_num - 1), column=8, value=row.iloc[2])

                if row.iloc[3] != prev_b10 or row.iloc[1] != prev_f10 or row.iloc[2] != prev_h10:
                    page_num = 1

                prev_b10 = row.iloc[3]
                prev_f10 = row.iloc[1]
                prev_h10 = row.iloc[2]

            copier_style_ligne(ws, 13, ligne)
            ws.cell(row=ligne, column=1, value=row.iloc[4])
            ws.cell(row=ligne, column=2, value=row.iloc[5])
            ws.cell(row=ligne, column=3, value=row.iloc[6])
            ws.cell(row=ligne, column=4, value=row.iloc[7])
            ws.cell(row=ligne, column=7, value=row.iloc[21])
            ws.cell(row=ligne, column=8, value=row.iloc[17])

            ligne += 1

        if ligne > 13 + 43 * (page_global_num - 1):  # Vérifie si des données ont été ajoutées
            copier_style_ligne(ws, 38, 38 + 43 * (page_global_num - 1))
            source_cell = ws.cell(row=38, column=2)
            target_cell = ws.cell(row=38 + 43 * (page_global_num - 1), column=2)
            if source_cell.has_style:
                target_cell._style = source_cell._style
            target_cell.value = source_cell.value
            ws.cell(row=40 + 43 * (page_global_num - 1), column=2, value=inspecteur)
            ws.cell(row=40 + 43 * (page_global_num - 1), column=6, value=membre)
            total_pages_filter = total_pages_dict[(prev_b10, prev_f10, prev_h10)]
            ws.cell(row=43 + 43 * (page_global_num - 1), column=4, value=f"{page_num} sur {total_pages_filter}")

        temp_excel_file_path = resource_path('OUTPUT/NF_temp.xlsx')
        wb.save(temp_excel_file_path)

        # Sauvegarde en PDF après la génération du fichier Excel temporaire
        save_selected_NF_as_pdf()

        messagebox.showinfo("ممتاز", "تم تصدير المحضر الجماعي بنجاح")

    except Exception as e:
        messagebox.showerror("Erreur", f"Une erreur s'est produite: {e}")

# Appel de la fonction telecharger_NF() au démarrage du script
if __name__ == "__main__":
    telecharger_NF()
